from pathlib import Path
from openpyxl import Workbook, load_workbook
import pandas as pd


def path_check(file_path: any):  # check if filepath is pathlib Path
    if not isinstance(file_path, Path):
        file_path = Path(file_path)
        print(f"Converting {file_path} to pathlib.Path")
    return file_path


def workbook_check(file_path: any, file_name: str):  # check if excel exists in filepath
    if not (file_path/file_name).exists():
        new_book = Workbook()
        new_book.save(file_path/file_name)
        print(f'Created {file_path/file_name}')


def sheet_check(file_path: any, file_name: str, new_sheet_name: any, create_if_not_exist: bool):
    wb = load_workbook(file_path/file_name)
    if new_sheet_name not in wb.sheetnames:
        if create_if_not_exist:
            wb.create_sheet(new_sheet_name)
            wb.save(file_path/file_name)
            print(f'Created {new_sheet_name} in {file_path/file_name}')
        else:
            print(f'{new_sheet_name} does not exist in {file_path/file_name}')


def add_df_to_excel(df: pd.DataFrame,
                    file_path: any,
                    file_name: str,
                    new_sheet_name: str,
                    if_sheet_exists: str,
                    create_if_not_exist: bool = True,
                    keep_index: bool = False):

    file_path = path_check(file_path)  # path format check
    workbook_check(file_path, file_name)   # workbook check
    sheet_check(file_path, file_name, new_sheet_name, create_if_not_exist)  # sheet check

    with pd.ExcelWriter(file_path/file_name, engine='openpyxl', mode='a', if_sheet_exists=if_sheet_exists) as writer:
        df.to_excel(writer, sheet_name=new_sheet_name, index=keep_index)
        print(f"{new_sheet_name} has been saved to {file_name} at {file_path}\n")
