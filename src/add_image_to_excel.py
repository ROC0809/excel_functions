from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from add_df_to_excel import path_check, workbook_check


def add_image_to_excel(position: str,
                       gap: int,
                       figsize: tuple,
                       dpi: int,
                       image_path: any,
                       image_name: str,
                       excel_path: any,
                       excel_name: str,
                       sheet_name: str):

    image_path = path_check(image_path)
    excel_path = path_check(excel_path)

    image_check(image_path, image_name)
    workbook_check(file_path=excel_path, file_name=excel_name)

    book = load_workbook(excel_path/excel_name)
    if sheet_name not in book.sheetnames:
        book.create_sheet(sheet_name)
        book.save(excel_path/excel_name)
        print(f"{sheet_name} created!")

    sheet = book[sheet_name]
    image_pos = locate_image_pos(sheet, position, gap)

    img = Image(image_path/image_name)
    img.width, img.height = figsize_to_WH(figsize, dpi)

    sheet.add_image(img, image_pos)
    book.save(excel_path/excel_name)
    print(f"\n{image_name} added to excel {excel_name}, sheet {sheet_name}")


def image_check(image_path: any, image_name: str):
    if not (image_path/image_name).exists():
        raise ValueError("Image path or image name is not correct")


def get_max_rows(ws):
    max_row_count = 0
    # iterate over all the columns to find the rows with the most data
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        current_row_count = len([cell for cell in ws[col_letter] if cell.value])
        if current_row_count >= max_row_count:
            max_row_count = current_row_count
    return max_row_count


def locate_image_pos(sheet, position, gap):
    if position not in ['right', 'bottom']:
        raise ValueError(f"{position} is not a valid input for position. Only 'right' and 'bottom' are accepted!")
    elif position == 'right':
        max_col = sheet.max_column
        target = max_col + gap
        max_col_excel_id = get_column_letter(target)
        image_pos = str(max_col_excel_id) + str(1)
    elif position == 'bottom':
        max_row = sheet.max_row
        target = max_row + gap
        col_excel_id = 'A'
        image_pos = col_excel_id + str(target)
    return image_pos


def figsize_to_WH(figsize, dpi):
    figsize_w, figsize_h = figsize
    img_w = figsize_w * dpi
    img_h = figsize_h * dpi
    return img_w, img_h
