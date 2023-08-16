import openpyxl.worksheet.worksheet
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from add_df_to_excel import add_df_to_excel


def save_locked_excel(
        target_df: pd.DataFrame,
        output_path: Path or str,
        output_name: str,
        new_sheet_name: str,
        excel_password: str = None,
        lock_columns: list = None,
        coloring_columns: dict = None,
        link_columns: list = None,
        validation_dict: dict = None,
        validation_kwargs: dict = None):

    """input check"""  # Check if validation_dict is provided, then validation_kwargs should also be provided
    if validation_dict is not None and validation_kwargs is None:
        raise ValueError("validation_dict is provided, "
                         "but validation_kwargs is None. "
                         "Both should be provided together.")

    """dataframe to excel with custom function"""
    add_df_to_excel(
        df=target_df,
        file_path=output_path,
        file_name=output_name,
        new_sheet_name=new_sheet_name,
        if_sheet_exists='replace',
        create_if_not_exist=True,
        keep_index=False
    )

    """load workbook"""
    workbook = load_workbook(output_path/output_name)
    worksheet = workbook[new_sheet_name]

    """unlock all cells in the worksheet"""
    # Explicitly unlock all cells in the worksheet due to openpyxl behaviors
    unlocked_protection = Protection(locked=False)
    for row in worksheet.iter_rows():
        for cell in row:
            cell.protection = unlocked_protection

    """lock columns"""
    if lock_columns:
        lock(
            target_df=target_df,
            locked_columns=lock_columns,
            worksheet=worksheet
        )
        print(f"locked columns: {lock_columns}")

    """fill background colors"""
    if coloring_columns:
        FF_coloring_columns = {
            col: hex2FF(color_code) for col, color_code in coloring_columns.items()
        }
        fill_background_color(
            target_df=target_df,
            coloring_columns=FF_coloring_columns,
            worksheet=worksheet
        )
        print(f"coloring columns: {coloring_columns}")

    """hyperlink columns"""
    if link_columns:
        link(
            target_df=target_df,
            link_columns=link_columns,
            worksheet=worksheet
        )
        print(f"hyperlinked columns: {link_columns}")

    """data validation"""
    if validation_dict:
        for col_name, validation_type in validation_dict.items():
            current_kwargs = validation_kwargs[validation_type]
            current_validation = create_validate(
                validate_type=validation_type, **current_kwargs
            )
            current_col_idx = target_df.columns.get_loc(col_name) + 1
            validate(
                col_idx=current_col_idx,
                target_df_len=len(target_df),
                validation=current_validation,
                worksheet=worksheet
            )
            print(f"validated column: {col_name} to {validation_type}")

    # Protect the worksheet
    worksheet.protection.sheet = True
    worksheet.protection.password = excel_password

    workbook.save(output_path/output_name)
    workbook.close()


def lock(target_df: pd.DataFrame,
         locked_columns: list,
         worksheet: openpyxl.worksheet.worksheet.Worksheet):
    # Lock specific columns
    locked_col_indices = [target_df.columns.get_loc(col) + 1 for col in locked_columns]
    locked_protection = Protection(locked=True)
    for col_idx in locked_col_indices:
        for cell in worksheet.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
            for c in cell:
                c.protection = locked_protection


def fill_background_color(target_df: pd.DataFrame,
                          coloring_columns: dict,
                          worksheet: openpyxl.worksheet.worksheet.Worksheet):
    # Fill background color in specific columns
    for col_name, color in coloring_columns.items():
        col_idx = target_df.columns.get_loc(col_name) + 1
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for cell in worksheet.iter_cols(min_col=col_idx, max_col=col_idx, min_row=1):
            for c in cell:
                c.fill = fill


def link(target_df: pd.DataFrame,
         link_columns: list,
         worksheet: openpyxl.worksheet.worksheet.Worksheet):
    for col_name in link_columns:
        col_idx = target_df.columns.get_loc(col_name) + 1
        for cell in worksheet.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2): # Skip header
            for c in cell:
                if c.value and isinstance(c.value, str) and c.value.startswith('http'):
                    c.hyperlink = c.value
                    c.value = c.value


def create_validate(validate_type, **kwargs):
    supported_validate_types = ['date', 'whole', 'textLength', 'time', 'custom', 'decimal', 'list']
    if validate_type not in supported_validate_types:
        raise ValueError(f"validate_type must be one of {supported_validate_types}")
    else:
        validation = DataValidation(
            type=validate_type, **kwargs
        )
    return validation


def validate(col_idx: int,
             target_df_len: int,
             validation: openpyxl.worksheet.datavalidation.DataValidation,
             worksheet: openpyxl.worksheet.worksheet.Worksheet):
    worksheet.add_data_validation(validation)
    col_letter = get_column_letter(col_idx)
    for row_idx in range(2, target_df_len + 2):  # Adjusting for 1-indexing and header row
        cell = worksheet[f'{col_letter}{row_idx}']
        validation.add(cell)


def hex2FF(color_code):
    if color_code.startswith("FF") and len(color_code) == 8:
        return color_code.upper()
    if color_code.startswith("#"):
        color_code = color_code[1:]
    # Add the alpha channel (FF) for full opacity
    return "FF" + color_code.upper()
